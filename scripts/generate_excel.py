#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
外汇牌价监控 - Excel生成模块
每个银行+币种一个sheet，自动覆盖历史文件
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 配置
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data"
HISTORY_FILE = DATA_DIR / "history.json"
EXCEL_FILE = BASE_DIR / "外汇牌价监控数据.xlsx"

# 样式定义
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DATA_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="微软雅黑", size=10, color="666666")

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# 银行颜色配置
BANK_COLORS = {
    "农业银行": {"header": "00753A", "light": "E8F5E9"},
    "民生银行": {"header": "005BAC", "light": "E3F2FD"},
    "中国银行": {"header": "C41E3A", "light": "FFEBEE"},
}


def load_history():
    """加载历史数据"""
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def create_summary_sheet(wb, history):
    """创建汇总sheet"""
    ws = wb.active
    ws.title = "实时汇总"
    
    # 标题
    ws.merge_cells("A1:F1")
    ws["A1"] = "外汇牌价实时监控汇总"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35
    
    # 更新时间
    ws.merge_cells("A2:F2")
    ws["A2"] = f"更新时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = CENTER_ALIGN
    ws.row_dimensions[2].height = 20
    
    # 表头
    headers = ["银行", "币种", "最新现汇买入价", "更新时间", "今日波动次数", "数据来源"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # 数据
    row = 5
    for key, records in sorted(history.items()):
        if not records:
            continue
        
        bank, currency = key.split("_", 1)
        latest = records[-1]
        
        # 计算今日波动次数
        today = datetime.now().strftime("%Y-%m-%d")
        today_changes = sum(1 for r in records if r["timestamp"].startswith(today))
        
        data = [
            bank,
            currency,
            float(latest["value"]) if latest["value"] else 0,
            latest["timestamp"],
            today_changes,
            latest.get("source", "未知"),
        ]
        
        bank_color = BANK_COLORS.get(bank, {"light": "F5F5F5"})
        fill = PatternFill(start_color=bank_color["light"], end_color=bank_color["light"], fill_type="solid")
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            cell.fill = fill
        
        row += 1
    
    # 调整列宽
    col_widths = [12, 10, 18, 22, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    return ws


def create_bank_currency_sheet(wb, bank, currency, records):
    """创建单个银行+币种的sheet"""
    sheet_name = f"{bank}-{currency}"
    ws = wb.create_sheet(title=sheet_name)
    
    bank_color = BANK_COLORS.get(bank, {"header": "4472C4", "light": "E8F0FE"})
    header_fill = PatternFill(start_color=bank_color["header"], end_color=bank_color["header"], fill_type="solid")
    light_fill = PatternFill(start_color=bank_color["light"], end_color=bank_color["light"], fill_type="solid")
    
    # 标题
    ws.merge_cells("A1:E1")
    ws["A1"] = f"{bank} - {currency} 现汇买入价历史"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 35
    
    # 统计信息
    if records:
        values = [float(r["value"]) for r in records if r["value"]]
        latest = records[-1]
        
        ws.merge_cells("A2:E2")
        ws["A2"] = f"最新价: {latest['value']} | 最高: {max(values):.3f} | 最低: {min(values):.3f} | 总记录数: {len(records)}"
        ws["A2"].font = SUBTITLE_FONT
        ws["A2"].alignment = CENTER_ALIGN
        ws.row_dimensions[2].height = 20
    
    # 表头
    headers = ["序号", "时间戳", "现汇买入价", "变动值", "数据来源"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # 数据（倒序，最新的在最上面）
    for i, record in enumerate(reversed(records), 1):
        row = 4 + i
        
        # 计算变动值
        old_value = record.get("old_value")
        new_value = record["value"]
        change = ""
        if old_value and new_value:
            try:
                diff = float(new_value) - float(old_value)
                change = f"{'+' if diff > 0 else ''}{diff:.3f}"
            except:
                change = ""
        
        data = [
            i,
            record["timestamp"],
            float(record["value"]) if record["value"] else 0,
            change,
            record.get("source", ""),
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            
            # 隔行变色
            if i % 2 == 0:
                cell.fill = light_fill
            
            # 涨跌颜色
            if col == 4 and change:
                if change.startswith("+"):
                    cell.font = Font(name="微软雅黑", size=10, color="C00000", bold=True)
                elif change.startswith("-"):
                    cell.font = Font(name="微软雅黑", size=10, color="00B050", bold=True)
    
    # 调整列宽
    col_widths = [8, 22, 15, 12, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 冻结首行
    ws.freeze_panes = "A5"
    
    return ws


def generate_excel():
    """生成Excel文件"""
    print("生成Excel文件...")
    
    history = load_history()
    
    if not history:
        print("  暂无历史数据，创建空Excel")
        history = {}
    
    wb = Workbook()
    
    # 1. 创建汇总sheet
    create_summary_sheet(wb, history)
    
    # 2. 为每个银行+币种创建sheet
    for key, records in sorted(history.items()):
        bank, currency = key.split("_", 1)
        if records:  # 只有有数据的才创建
            create_bank_currency_sheet(wb, bank, currency, records)
    
    # 3. 保存文件（覆盖）
    wb.save(EXCEL_FILE)
    print(f"  ✓ Excel已保存: {EXCEL_FILE}")
    print(f"  ✓ 共 {len(history)} 个sheet")
    
    return EXCEL_FILE


if __name__ == "__main__":
    generate_excel()
